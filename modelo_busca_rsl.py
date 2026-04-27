# Importa o módulo sys para interagir com o sistema, útil para finalizar o script em caso de erro
import sys
# Importa o módulo os para interagir com o sistema operacional, lidar com caminhos de arquivos etc
import os
# Importa o módulo subprocess para permitir a execução de comandos no terminal (usado para instalar pacotes)
import subprocess
# Importa o módulo urllib.parse para codificar as URLs nas buscas das APIs
import urllib.parse
# Importa o módulo difflib para o cálculo de similaridade de strings (Auditoria)
import difflib
# Importa recursos para ignorar alertas chatos de SSL ao testar links quebrados
import warnings
from urllib3.exceptions import InsecureRequestWarning
warnings.simplefilter('ignore', InsecureRequestWarning)

# ==========================================
# Instalador automático de dependências
# ==========================================
# Define a função que verifica e instala as bibliotecas necessárias para o script rodar
def install_requirements():
    # Cria um dicionário mapeando o nome que a gente usa no import para o nome do pacote no pip
    required_packages = {
        'pyzotero': 'pyzotero',
        'pandas': 'pandas',
        'requests': 'requests',
        'urllib3': 'urllib3',
        'pdfplumber': 'pdfplumber',
        'bs4': 'beautifulsoup4',
        'openpyxl': 'openpyxl'
    }
    
    # Cria uma lista vazia para guardar os nomes dos pacotes que ainda não estão instalados
    missing_packages = []
    # Percorre o dicionário testando cada pacote um por um
    for module_name, pip_name in required_packages.items():
        # Tenta importar o módulo
        try:
            # Importa de forma dinâmica usando a função embutida __import__
            __import__(module_name)
        # Se der erro de importação, significa que o pacote não está instalado
        except ImportError:
            # Adiciona o nome do pacote na lista de pacotes ausentes
            missing_packages.append(pip_name)
            
    # Se a lista de pacotes ausentes não estiver vazia, ou seja, falta instalar algo
    if missing_packages:
        # Mostra na tela uma mensagem avisando quais pacotes vão ser instalados
        print(f"📦 Dependências ausentes encontradas. Instalando: {', '.join(missing_packages)}...")
        # Tenta executar o comando de instalação
        try:
            # Roda no terminal o pip install com a lista de pacotes que faltam
            subprocess.check_call([sys.executable, "-m", "pip", "install", *missing_packages])
            # Avisa que a instalação deu certo e o script vai continuar
            print("✅ Dependências instaladas com sucesso! Reiniciando carregamento...\n")
        # Caso ocorra algum erro inesperado durante a instalação
        except Exception as e:
            # Mostra qual foi o erro que aconteceu
            print(f"❌ Erro ao instalar dependências: {e}")
            # Avisa o usuário como ele pode tentar instalar manualmente no terminal
            print(f"⚠️ Por favor, instale manualmente rodando: pip install {' '.join(missing_packages)}")
            # Encerra o programa com código de erro 1
            sys.exit(1)

# Chama a função para garantir que tudo está instalado antes de rodar o resto do código
install_requirements()

# Após garantir que todos os pacotes existem, fazemos as importações principais para o funcionamento
# Importa o Zotero da biblioteca pyzotero, usado para conectar na conta e buscar os artigos
from pyzotero import zotero
# Importa a biblioteca pandas com o apelido pd, usada para manipular dados em formato de tabela
import pandas as pd
# Importa a biblioteca re, que permite usar expressões regulares para buscar padrões em textos (como datas e links)
import re
# Importa difflib para comparar similaridade entre textos (como títulos)
import difflib
# Importa a biblioteca time, útil para pausar o script e evitar banimento em APIs
import time
# Importa a biblioteca requests, essencial para fazer chamadas HTTP na internet (baixar páginas, falar com APIs)
import requests
# Importa urllib3, que lida com requisições HTTP em baixo nível, usada aqui pra desligar alguns avisos
import urllib3
# Importa o módulo io, que permite lidar com dados em memória como se fossem arquivos reais
import io
# Importa o pdfplumber, que serve para ler o conteúdo de PDFs com melhor precisão de espaços e layout
import pdfplumber
# Importa o BeautifulSoup do bs4, ferramenta muito boa para extrair dados de páginas HTML da web
from bs4 import BeautifulSoup
# Importa o Workbook do openpyxl, que permite criar um novo arquivo de planilha do Excel
from openpyxl import Workbook
# Importa os módulos de estilo do openpyxl para deixar a planilha bonita (cores, fontes, bordas e alinhamentos)
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# Importa a função que transforma as tabelas do pandas diretamente em linhas do Excel
from openpyxl.utils.dataframe import dataframe_to_rows

# Força o sistema a usar codificação UTF-8 no Windows, evitando erros quando tentamos imprimir emojis no terminal
if sys.platform.startswith('win'):
    # Tenta reconfigurar a saída padrão para aceitar caracteres especiais
    try:
        # Aplica o utf-8 no stdout
        sys.stdout.reconfigure(encoding='utf-8')
    # Se a versão do Python não suportar ou der problema, simplesmente ignora e segue em frente
    except AttributeError:
        # Passa reto caso não consiga
        pass

# Desativa os avisos chatos do urllib3 sobre requisições inseguras quando usamos verify=False nos requests
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ==========================================
# Configurações da automação da Revisão Sistemática da Literatura
# ==========================================
# O código de usuário exclusivo do Zotero para o script saber de qual biblioteca baixar os artigos
ID_ZOTERO = ''
# Sua senha de acesso (API Key) para provar que você autorizou o script a ler seus dados
CHAVE_API = ''
# E-mail verdadeiro para entrar na "fila educada" das APIs e evitar bloqueios nas buscas
SEU_EMAIL = ''

# Configuração para padronizar os números do Ranking com zero à esquerda
DIGITOS_RANK = 2
# Pede o nome do usuário para personalizar a saída
nome_usuario = input("Digite seu nome para o relatório: ")
# ==========================================

# Cria uma instância da conexão com o Zotero, passando o id, que é uma biblioteca de usuário ('user'), e a chave de acesso
zot = zotero.Zotero(ID_ZOTERO, 'user', CHAVE_API)

# --- Funções utilitárias e interface do terminal ---
# Define uma função simples para limpar textos que podem vir sujos da internet
def limpar_texto(texto):
    # Se o texto for vazio ou for um valor nulo do pandas, retorna uma string vazia
    if not texto or pd.isna(texto): return ""
    # Remove espaços duplicados e apaga qualquer coisa que pareça uma tag HTML usando expressão regular
    txt = re.sub(r'\s+', ' ', re.sub('<[^<]+?>', '', str(texto))).strip()
    
    # Vacina 1: Tenta remover sujeiras de prefixos do tipo "Nome da Revista - O Resumo Começa Aqui"
    m_prefix = re.match(r'^([A-Z][a-zA-Z0-9\s\&,\.\:\-]{3,60}?)\s+-\s+(.*)', txt)
    if m_prefix: txt = m_prefix.group(2)
    
    # Vacina 2: Remove a palavra Abstract ou Resumo que fica sobrando no começo do texto
    txt = re.sub(r'(?i)^(abstract|resumo|summary|resumen)[\s:\-]+', '', txt).strip()
    
    return txt

# Define uma função para encontrar o DOI de um artigo dentro de um texto qualquer
def extrator_doi(texto):
    # Se não tiver texto, não tem como extrair, retorna vazio
    if not texto: return ""
    # Procura pelo padrão clássico de um DOI (que sempre começa com 10.algumacoisa) usando regex, ignorando maiúsculas e minúsculas
    m = re.search(r'(10\.\d{4,9}/[-._;()/:A-Z0-9a-z]+)', str(texto).replace('doi.org/', ''), re.IGNORECASE)
    # Se achou algo, retorna o pedaço exato do texto, senão retorna vazio
    return m.group(1) if m else ""

# Define uma função para capturar apenas o ano no meio de uma data completa ou de um texto bagunçado
def extrator_ano(texto):
    # Retorna vazio se não tiver texto de entrada
    if not texto: return ""
    # Busca um número de 4 dígitos que comece com 19 ou 20, representando anos do século 20 e 21
    m = re.search(r'\b(19\d{2}|20\d{2})\b', str(texto))
    # Retorna o ano que foi encontrado, ou então devolve os primeiros 4 caracteres do texto como plano B
    return m.group(1) if m else str(texto)[:4]

# Define uma função para comparar se dois títulos são do mesmo artigo
def verificar_similaridade_titulo(t1, t2):
    if not t1 or not t2: return False
    # Limpa pontuação e deixa tudo minúsculo para a comparação ser justa
    t1_c = re.sub(r'[^\w\s]', '', str(t1).lower()).strip()
    t2_c = re.sub(r'[^\w\s]', '', str(t2).lower()).strip()
    if not t1_c or not t2_c: return False
    # Se um título está perfeitamente contido no outro, é o mesmo artigo
    if t1_c in t2_c or t2_c in t1_c: return True
# Avalia a taxa de similaridade (acima de 75% é considerado o mesmo artigo)
    return difflib.SequenceMatcher(None, t1_c, t2_c).ratio() > 0.75

# Define uma função para limpar o lixo de cabeçalho (nomes, orientadores) do início do resumo
def limpar_cabecalho_resumo(texto_abstract, titulo_conhecido, autores_conhecidos):
    linhas = str(texto_abstract).split('\n')
    linhas_validas = []
    capturando = False
    
    palavras_descarte = ['supervised by', 'bachelor of', 'master of', 'university', 'department of', 'submitted to', 'faculty of', 'approved by', 'thesis']
    
    for linha in linhas:
        linha_str = linha.strip()
        if not linha_str: continue
            
        linha_lower = linha_str.lower()
        
        if not capturando:
            if any(p in linha_lower for p in palavras_descarte):
                continue
            
            if titulo_conhecido and (linha_lower in titulo_conhecido.lower() or titulo_conhecido.lower() in linha_lower):
                continue
                
            if autores_conhecidos and len(linha_str) < 50:
                is_autor = False
                for aut in str(autores_conhecidos).split(';'):
                    parte_nome = aut.split(',')[0].strip().lower()
                    if parte_nome and parte_nome in linha_lower:
                        is_autor = True
                        break
                if is_autor: continue
            
            if len(linha_str.split()) > 5:
                capturando = True
                linhas_validas.append(linha_str)
        else:
            linhas_validas.append(linha_str)
            
    texto_limpo = " ".join(linhas_validas)
    if len(texto_limpo) > 100:
        return texto_limpo
    return str(texto_abstract).replace('\n', ' ')

# Define uma função que organiza as informações encontradas num painel bonito para mostrar no terminal
def imprimir_resumo_artigo(titulo, estado, fontes, idx, total):
    # Pula uma linha e escreve qual é o número do artigo atual comparado ao total, além do começo do título dele
    print(f"\n📖 [{idx}/{total}] Artigo: {str(titulo)[:100]}...")
    
    # Cria uma linha separadora com sinais de mais e menos para formar a tabela de texto
    hr = "+" + "-"*15 + "+" + "-"*20 + "+" + "-"*30 + "+"
    # Imprime a linha de cima da tabela
    print(hr)
    # Imprime o cabeçalho das colunas informando os espaços alinhados à esquerda
    print(f"| {'CAMPO':<13} | {'STATUS':<18} | {'FONTE':<28} |")
    # Imprime a linha embaixo do cabeçalho
    print(hr)
    
    # Cria um dicionário que liga o nome bonito da coluna à chave onde o valor real foi salvo no programa
    mapeamento = {
        'DOI': 'doi', 'Resumo': 'resumo', 'Título': 'titulo', 
        'Autores': 'autores', 'Ano': 'ano', 'Link': 'link'
    }
    
    # Percorre cada campo para verificar se foi encontrado e imprimir o resultado
    for c, key in mapeamento.items():
        # Pega o valor daquela informação (por exemplo, o texto do resumo) ou vazio se não tiver
        val = estado.get(key, '')
        # Pega o nome do lugar onde aquela informação foi encontrada, ou N/A se ninguém achou
        fonte = fontes.get(c, 'N/A')
        
        # Decide se o campo foi realmente achado verificando se ele não está vazio e não é a string N/A
        found = val and str(val).strip() != "" and val != 'N/A'
        # Define o texto do status baseado na verificação anterior
        status_text = "Encontrado" if found else "Não Encontrado"
        # Escolhe qual emoji colocar do lado (um verde ou um vermelho)
        icone = "✅" if found else "❌"
        
        # O terminal do Windows se confunde com os espaços dos emojis, então a gente calcula pra não entortar as linhas da tabela
        espacos_status = 16 if found else 15
        # Imprime a linha da tabela preenchida com a informação deste campo, o ícone e onde ele foi achado
        print(f"| {c:<13} | {icone} {status_text:<{espacos_status}} | {fonte:<28} |")
        
    # Imprime a linha de baixo para fechar a tabela bonitinha
    print(hr)

# Define a função principal de busca na web para quando falta resumo, tentaremos extrair diretamente do PDF ou da página do artigo
def extrair_metadados_web_pdf(url, titulo_conhecido="", autores_conhecidos=""):
    # Documentação da função descrevendo o que ela tenta fazer
    """Tenta baixar e ler os metadados diretamente do PDF ou de Scrapers HTML"""
    # Inicializa um dicionário padrão só com valores vazios que vamos tentar preencher
    retorno = {'resumo': '', 'titulo': '', 'autores': '', 'ano': '', 'fonte': '', 'doi': ''}
    
    # Tenta lidar com links específicos do Figshare porque eles bloqueiam robôs na página principal
    m_figshare = re.search(r'figshare\.com/articles/[^/]+/[^/]+/(\d+)', str(url))
    # Se a URL fornecida bater com a lógica do Figshare, ativamos essa condição
    if m_figshare:
        # Usa bloco try para evitar travamentos caso a internet caia no meio do processo
        try:
            # Faz um pedido para a API secreta do Figshare usando o código do artigo capturado na URL
            r_api = requests.get(f"https://api.figshare.com/v2/articles/{m_figshare.group(1)}", timeout=10)
            # Confere se a resposta do servidor deu tudo certo (código 200)
            if r_api.status_code == 200:
                # Transforma o texto de resposta em um dicionário de dados (JSON)
                data = r_api.json()
                # Puxa a descrição principal, que geralmente serve como resumo
                desc = limpar_texto(data.get("description", ""))
                
                # Prepara uma lista vazia para agrupar os autores
                autores = []
                # Passa por cada autor no JSON retornado
                for a in data.get("authors", []):
                    # Pega o sobrenome
                    ln = a.get("last_name", "")
                    # Pega o primeiro nome
                    fn = a.get("first_name", "")
                    # Se tiver os dois nomes, junta com vírgula (padrão de citação) e adiciona na lista
                    if ln and fn: autores.append(f"{ln}, {fn}")
                    # Caso contrário, pega só o nome completo e adiciona
                    else: autores.append(a.get("full_name", ""))
                # Transforma a lista de autores em uma string única dividida pela palavra and
                str_autores = " and ".join(autores)
                
                # Extrai e limpa o título da pesquisa
                titulo = limpar_texto(data.get("title", ""))
                
                # Cria variáveis para capturar dados extras, como o grau da pesquisa e o orientador
                qualificacao = ""
                orientador = ""
                # Varre os campos customizados porque tese e dissertação jogam esses dados lá
                for field in data.get("custom_fields", []):
                    # Se o nome do campo for Qualification Name, salva o valor nele
                    if field.get("name") == "Qualification Name":
                        qualificacao = field.get("value", "")
                    # Se for Supervisors, guarda o nome do orientador
                    if field.get("name") == "Supervisors":
                        orientador = field.get("value", "")
                
                # Pega a data completa em que foi publicado
                data_pub = data.get("published_date", "")
                str_data = ""
                ano = ""
                # Verifica se a data foi retornada
                if data_pub:
                    # Importa localmente o pacote para tratar datas
                    from datetime import datetime
                    # Tenta converter o formato da data para tirar o ano de forma segura
                    try:
                        # Analisa o texto até o décimo caractere que compõe a data padrão YYYY-MM-DD
                        dt = datetime.strptime(data_pub[:10], "%Y-%m-%d")
                        # Formata uma versão bonita com mês e ano
                        str_data = dt.strftime("%B, %Y")
                        # Formata só o ano
                        ano = dt.strftime("%Y")
                    # Se a data vier estranha, roda o plano de escape
                    except:
                        # Guarda a string original cortada
                        str_data = data_pub[:10]
                        # Tenta usar nossa função de extração de ano
                        ano = extrator_ano(data_pub)
                
                # Pega o DOI retornado pela API do Figshare se houver
                doi = data.get("doi", "")
                
                # Prepara um aglomerado para juntar esses metadados soltos da tese
                partes = []
                # Se tem autores, joga na lista
                if str_autores: partes.append(str_autores)
                # Se tem título, bota entre aspas na lista
                if titulo: partes.append(f'"{titulo}"')
                # Se tem a qualificação do curso, adiciona também
                if qualificacao: partes.append(qualificacao)
                # Adiciona a data formatada
                if str_data: partes.append(str_data)
                # Coloca o orientador
                if orientador: partes.append(orientador)
                
                # Concatena tudo com vírgulas pra fazer uma referência bibliográfica no estilo citação
                texto_completo = ", ".join(partes)
                # Se formou a citação, bota um ponto e quebra de linha antes de botar o resumo do estudo
                if texto_completo: texto_completo += ".\n" + desc
                # Senão, o texto final é só a descrição sozinha mesmo
                else: texto_completo = desc
                
                # Se gerou um texto com tudo, atualiza o dicionário de retorno pra gente não perder viagem
                if texto_completo:
                    retorno['resumo'] = texto_completo
                    retorno['titulo'] = titulo
                    retorno['autores'] = str_autores
                    retorno['ano'] = ano
                    retorno['doi'] = doi
                    retorno['fonte'] = "API Figshare"
                    # Já encerra a função e retorna os dados aqui mesmo
                    return retorno
        # Se deu algum erro no meio (a API falhou), passa reto
        except: pass

    # Inicia o fluxo normal para baixar o arquivo se não for figshare ou se o figshare falhou
    # Criamos um cabeçalho de requisição para imitar um navegador real (Google Chrome) assim evitamos bloqueios fáceis de firewalls
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'application/pdf,application/xhtml+xml,text/html;q=0.9,*/*;q=0.8',
        'Accept-Language': 'pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7'
    }
    
    # Inicia o bloco de tentativa principal de conexão na web
    try:
        # Pede para baixar o link. Aumentamos o timeout para (10, 30) para tolerar repositórios institucionais lentos (como UFRN)
        response = requests.get(url, headers=headers, timeout=(10, 30), verify=False, allow_redirects=True)
        # Se a página ou PDF existirem, o código será 200 de sucesso
        if response.status_code == 200:
            # Observa o começo do arquivo pra descobrir se é um arquivo de formato PDF (olhamos 100 bytes para tolerar cabeçalhos sujos)
            if b'%PDF' in response.content[:100]:
                # Transforma o conteúdo baixado num arquivo em memória
                f = io.BytesIO(response.content)
                # Usa a biblioteca pdfplumber pra ler o conteúdo sem grudar palavras
                with pdfplumber.open(f) as pdf:
                    # Tenta aproveitar as informações embutidas nas propriedades ocultas do PDF
                    if pdf.metadata:
                        retorno['titulo'] = pdf.metadata.get('Title', '')
                        retorno['autores'] = pdf.metadata.get('Author', '')
                    
                    # Cria uma variável vazia de texto para jogar todo o texto que lermos do arquivo
                    total_text = ""
                    # Faz um loop para ler, no máximo, as 30 primeiras páginas do PDF
                    for i in range(min(30, len(pdf.pages))):
                        # Extrai o texto da página específica da vez preservando o layout base
                        page_text = pdf.pages[i].extract_text()
                        # Se vier texto válido, soma com o que já foi lido pulando uma linha
                        if page_text: total_text += "\n" + page_text
                
                # Regex refinada com stop-words rígidas (contents, glossary, index, etc) para evitar vazamento
                regex_resumo = r'(?i)\b(?:abstract|resumo|resumen|summary)\b[\s\n\r:]+([\s\S]{100,4500}?)(?=\b(?:contents|glossary|table of contents|index|introduction|introdução|introduccion|1\. introduction|keywords|palavras-chave|palabras clave|index terms|acknowledgment|chapter 1|1\.)\b|\Z)'
                # Aplica o super filtro no texto gigante que acabamos de ler
                match = re.search(regex_resumo, total_text)
                
                # Se o filtro casou perfeitamente com o texto e o resumo foi isolado
                if match:
                    # Aplica a heurística de limpeza de cabeçalho para remover metadados e nomes
                    titulo_para_limpeza = titulo_conhecido or retorno['titulo']
                    autores_para_limpeza = autores_conhecidos or retorno['autores']
                    texto_isolado = limpar_cabecalho_resumo(match.group(1).strip(), titulo_para_limpeza, autores_para_limpeza)
                    
                    # Limpa o texto isolado pra remover tags soltas e garantir que está limpo
                    retorno['resumo'] = limpar_texto(texto_isolado)
                    # Avisa que a fonte dessa informação foi nosso leitor com regex
                    retorno['fonte'] = "Leitor PDF (pdfplumber + Heurística)"
                    # Termina a função entregando esses dados
                    return retorno
                # Se não conseguiu isolar com a regra e falhou em ler exatamente a delimitação
                else:
                    # Pega tudo que leu e só remove sujeiras das pontas
                    txt_bruto = limpar_texto(total_text.replace('\n', ' '))
                    # Verifica se o arquivo tem pelo menos bastante texto (mais de mil letras)
                    if len(txt_bruto) > 1000:
                        # Captura um pedaço do começo na força bruta e anexa um aviso para a gente revisar depois na planilha
                        retorno['resumo'] = txt_bruto[1000:2500] + "... [Texto Forçado - Revise]"
                        # Registra que foi no desespero que pegou o texto
                        retorno['fonte'] = "PDF (Força Bruta)"
                        # Envia os dados pra trás
                        return retorno
            
            # Se não tem a marca do PDF, então só pode ser uma página HTML de um repositório da web
            elif b'%PDF' not in response.content[:100]:
                # Transforma a resposta no código fonte HTML legível
                html = response.text
                # Invoca o BeautifulSoup pra mastigar o HTML pra gente e virar algo fácil de pesquisar
                soup = BeautifulSoup(html, 'html.parser')
                
                # Passa por todas as marcações chamadas meta (elas carregam as informações ocultas dos sites)
                for tag in soup.find_all('meta'):
                    # Pega o atributo nome da tag (deixando minúsculo)
                    nome = str(tag.get('name', '')).lower()
                    # Pega o atributo de propriedade (usado muito no protocolo Open Graph de redes sociais)
                    prop = str(tag.get('property', '')).lower()
                    # Pega o conteúdo de fato onde tá o texto valioso
                    conteudo = limpar_texto(tag.get('content', ''))
                    
                    # Se não tem conteúdo na tag, pula pra próxima pra não dar erro
                    if not conteudo: continue
                        
                    # Testa se o nome ou a propriedade sugerem que é aqui que fica o resumo do artigo (abstract, description etc)
                    if 'abstract' in nome or 'description' in prop or 'description' in nome or 'dc.description' in nome:
                        # Vacina Springer: Algumas editoras colocam o link do artigo dentro da tag description. Rejeita URLs e strings sem espaços.
                        if conteudo.startswith('http') or ' ' not in conteudo:
                            continue
                            
                        # Se o texto for razoavelmente grande e a gente não tiver achado o resumo ainda, guarda
                        if len(conteudo) > 100 and not retorno['resumo']: 
                            # Identifica se a página cortou o resumo no meio usando os pontinhos e evita pegar como principal
                            if len(conteudo) < 400 and (conteudo.strip().endswith('...') or conteudo.strip().endswith('…')):
                                retorno['resumo_truncado'] = conteudo
                            else:
                                retorno['resumo'] = conteudo
                    # Testa se o nome indica título original (usando tags comuns de bibliotecas online)
                    elif nome in ['citation_title', 'dc.title'] and not retorno['titulo']:
                        # Salva o conteúdo do título
                        retorno['titulo'] = conteudo
                    # Testa se a tag armazena os autores do artigo
                    elif nome in ['citation_author', 'dc.creator']:
                        # Se já tem algum autor guardado, anexa esse com um ponto e vírgula separando
                        if retorno['autores']: retorno['autores'] += f"; {conteudo}"
                        # Se não tinha ninguém, ele é o primeiro
                        else: retorno['autores'] = conteudo
                    # Testa as variações comuns para a data de publicação original do trabalho acadêmico
                    elif nome in ['citation_publication_date', 'citation_date', 'dc.date'] and not retorno['ano']:
                        # Passa pela nossa função de arrancar ano antes de salvar
                        retorno['ano'] = extrator_ano(conteudo)
                    # Tenta capturar o DOI das tags
                    elif nome in ['citation_doi', 'dc.identifier'] and not retorno['doi']:
                        if '10.' in conteudo: retorno['doi'] = extrator_doi(conteudo)
                
                # Vacina 3: Procura fisicamente no corpo da página por tags (div, section, p) que sejam marcadas como abstract (Resolve Code4Lib e similares)
                for tag in soup.find_all(re.compile('^(div|section|p)$')):
                    classes = " ".join(tag.get('class', [])).lower()
                    id_attr = str(tag.get('id', '')).lower()
                    
                    if 'abstract' in classes or 'abstract' in id_attr or 'summary' in classes:
                        txt_body = limpar_texto(tag.text)
                        # Só aceita se for razoavelmente grande e parecer um resumo de verdade
                        if len(txt_body) > 150:
                            retorno['resumo'] = txt_body
                            retorno['fonte'] = "Scraper HTML (Tag de Abstract)"
                            return retorno
                            
                # Depois de varrer todas as tags e a busca aprofundada, se encontrou o resumo inteiro na meta tag
                if retorno['resumo']:
                    # Atribui a fonte de HTML Meta Tags pro pessoal saber
                    retorno['fonte'] = "Scraper HTML (Meta Tags)"
                    # E devolve o relatório
                    return retorno
                
                # Se nada acima funcionou e a gente só tem aquele resumo cortado com "...", usamos ele como consolação
                if retorno.get('resumo_truncado'):
                    retorno['resumo'] = retorno['resumo_truncado']
                    retorno['fonte'] = "Scraper HTML (Meta Truncada)"
                    return retorno
                    
                # Alguns sites como a IEEE tem um formato diferente, vamos tentar uma pesquisa super específica se o link for de lá
                if 'ieee.org' in url:
                    # Procura na marra pelo texto que vem depois da palavra abstract na página toda
                    m = re.search(r'"abstract":"(.*?)"', html)
                    # Se encontrar o texto dentro dessa formatação estranha do IEEE
                    if m: 
                        # Limpa esse texto que ele pegou
                        retorno['resumo'] = limpar_texto(m.group(1))
                        # Coloca a fonte pra identificar
                        retorno['fonte'] = "Scraper HTML (IEEE)"
                        # Devolve
                        return retorno
                        
                # A ProQuest esconde o PDF, mas o resumo fica disponível numa div específica do HTML
                if 'proquest.com' in url:
                    div_abstract = soup.find(class_='abstract_Text')
                    if div_abstract:
                        retorno['resumo'] = limpar_texto(div_abstract.text)
                        retorno['fonte'] = "Scraper HTML (ProQuest)"
                        return retorno
                        
    # Se der erro de timeout ou conexão ser recusada, o try joga pra esse except pra não travar o programa
    except Exception as e:
        # Coloca o estado da fonte como erro, assim a gente sabe que não foi porque o artigo não tinha as informações, foi erro técnico
        retorno['fonte'] = f"Erro no Download"
    
    # Se o script tentou tudo acima e ainda não salvou nada de resumo na variável
    if not retorno['resumo']: retorno['fonte'] = "Não Encontrado"
    # Devolve esse dicionário que no fim pode ter encontrado o resumo por outros meios e atualizado partes dele
    return retorno

# --- Motor de automação RSL (Enriquecimento hierárquico) ---
# Função que lê um artigo do Zotero e vai caçando nos outros lugares as coisas que estão faltando
def buscar_metadados_hierarquia(item_zotero, idx, total, tem_ranking_manual=True):
    # Pega os dados brutos do Zotero que enviamos dentro da função
    d = item_zotero.get('data', {})
    
    # Captura o campo extra do zotero que guarda a classificação e limpa pra pegar só os números do começo antes de um espaço ou quebra de linha
    campo_extra = str(d.get('extra', ''))
    rank_match = re.search(r'^\s*(\d+)(?:\s|\n|$)', campo_extra)
    
    if rank_match:
        # Formata com a quantidade de zeros configurada
        rank_val = rank_match.group(1).zfill(DIGITOS_RANK)
    else:
        # Se houver rankings em algum outro item, joga pro final com 99, senão conta naturalmente (01, 02...)
        rank_val = ("9" * DIGITOS_RANK) if tem_ranking_manual else str(idx).zfill(DIGITOS_RANK)
    
    # Monta o dicionário de estado, que é onde deixamos os valores atuais para ir preenchendo os buracos
    estado = {
        # Guarda a classificação extraída
        'rank': rank_val,
        # O doi é lido do campo doi, chamando a função limpadora de lixo antes
        'doi': extrator_doi(d.get('DOI', '')),
        # Pega o título diretamente
        'titulo': d.get('title', ''),
        # O resumo do zotero vem do abstractNote, passa pelo limpador pra não vir caracteres errados
        'resumo': limpar_texto(d.get('abstractNote', '')),
        # Pega a data e a função corta só o ano
        'ano': extrator_ano(d.get('date', '')),
        # Junta todos os autores numa lista gigante separada por pontos e vírgulas (puxando o sobrenome primeiro e depois o primeiro nome)
        'autores': "; ".join([f"{a.get('lastName')}, {a.get('firstName')}" if a.get('lastName') and a.get('firstName') else a.get('lastName') or a.get('name', '') for a in d.get('creators',[])]),
        # Captura o site que estava cadastrado no zotero
        'link': d.get('url', '')
    }
    
    # Monta um dicionário para carregar de onde saiu a ideia do dado, a princípio, ou vieram do Zotero ou estão N/A (Não disponíveis)
    fontes = {
        'DOI': 'Zotero' if estado['doi'] else 'N/A',
        'Resumo': 'Zotero' if estado['resumo'] else 'N/A',
        'Título': 'Zotero' if estado['titulo'] else 'N/A',
        'Autores': 'Zotero' if estado['autores'] else 'N/A',
        'Ano': 'Zotero' if estado['ano'] else 'N/A',
        'Link': 'Zotero' if estado['link'] else 'N/A'
    }

    # Define uma função de verificação muito rápida que checa se todos os 4 campos vitais já estão preenchidos ou não
    def precisa_enriquecer():
        # Retorna verdadeiro se pelo menos um desses quatro campos ainda estiver vazio
        return not estado['resumo'] or not estado['doi'] or not estado['titulo'] or not estado['autores']

    # Prepara um cabeçalho pra acessar as APIs mostrando o email que a gente configurou pra ser legal com os servidores
    headers_padrao = {'User-Agent': f'Script_RSL/1.1 (mailto:{SEU_EMAIL})'}
    # Protege o campo do título, se não tiver a gente manda Unknown Title para não quebrar a busca
    titulo_query = estado['titulo'] or "Unknown Title"

    # 1. Busca na API oficial do Crossref
    # Se os campos vitais precisam de enriquecimento
    if precisa_enriquecer():
        # Usa try pra evitar erros da internet
        try:
            # Faz a busca passando o título do artigo no meio do link usando urllib para traduzir espaços
            r = requests.get(f"https://api.crossref.org/works?query.title={urllib.parse.quote(titulo_query)}&rows=1", headers=headers_padrao, timeout=5)
            # Se deu certo e o servidor retornou os itens pesquisados
            if r.status_code == 200 and r.json().get('message', {}).get('items'):
                # Pega o primeiro item (o mais correspondente com nosso título)
                it = r.json()['message']['items'][0]
                crossref_title = it.get('title', [''])[0] if it.get('title') else ''
                
                # Só aceita usar os dados se o título for realmente do mesmo artigo
                if verificar_similaridade_titulo(titulo_query, crossref_title):
                    # Se nós não tínhamos o DOI e o Crossref informou um
                    if not estado['doi'] and it.get('DOI'):
                        # Atualiza com o extrator
                        estado['doi'] = extrator_doi(it.get('DOI'))
                        # Atualiza a fonte
                        fontes['DOI'] = 'Crossref'
                    # Se não temos resumo e eles tem o abstract
                    if not estado['resumo'] and it.get('abstract'):
                        # Atualiza limpando o texto
                        estado['resumo'] = limpar_texto(it.get('abstract'))
                        # Atualiza a fonte do resumo
                        fontes['Resumo'] = 'Crossref'
                    # Se falta o ano e na devolução existe a parte de datas
                    if not estado['ano'] and it.get('issued', {}).get('date-parts'):
                        # Captura o primeiro digito da tabela aninhada de datas
                        estado['ano'] = str(it.get('issued')['date-parts'][0][0])
                        # Atualiza a fonte do ano
                        fontes['Ano'] = 'Crossref'
                    # Se não temos autores, tenta pegar do Crossref
                    if not estado['autores'] and it.get('author'):
                        lista_autores = [f"{a.get('family', '')}, {a.get('given', '')}".strip(', ') for a in it.get('author', []) if a.get('family')]
                        if lista_autores:
                            estado['autores'] = "; ".join(lista_autores)
                            if len(lista_autores) > 5:
                                estado['autores'] += " [Consórcio?]"
                            fontes['Autores'] = 'Crossref'
        # Em caso de falha não precisa gritar, só avança pra próxima tática
        except: pass

    # 2. Busca pela API do Semantic Scholar
    # Se ainda faltam coisas para enriquecer, tenta mais essa
    if precisa_enriquecer():
        # Abre try de precaução
        try:
            # Requisita na API do Semantic os campos: abstract, título, autores e ano. Só limitamos a 1 item pra vir os dados que importam
            r = requests.get(f"https://api.semanticscholar.org/graph/v1/paper/search?query={urllib.parse.quote(titulo_query[:100])}&limit=1&fields=abstract,title,authors,year", timeout=5)
            # Se funcionou e tem o dado
            if r.status_code == 200 and r.json().get('data'):
                # Salva o objeto na variável menor
                it = r.json()['data'][0]
                semantic_title = it.get('title', '')
                
                # Só processa se for do mesmo artigo
                if verificar_similaridade_titulo(titulo_query, semantic_title):
                    # Verifica o resumo se estamos precisando dele
                    if not estado['resumo'] and it.get('abstract'):
                        # Limpa e salva no estado
                        estado['resumo'] = limpar_texto(it.get('abstract'))
                        # Anota que veio daqui
                        fontes['Resumo'] = 'Semantic Scholar'
                    # Analisa o ano
                    if not estado['ano'] and it.get('year'):
                        # Salva como string do estado
                        estado['ano'] = str(it.get('year'))
                        # Anota o mérito
                        fontes['Ano'] = 'Semantic Scholar'
                    # Analisa os autores
                    if not estado['autores'] and it.get('authors'):
                        lista_autores = [a.get('name', '') for a in it.get('authors', []) if a.get('name')]
                        if lista_autores:
                            estado['autores'] = "; ".join(lista_autores)
                            if len(lista_autores) > 5:
                                estado['autores'] += " [Consórcio?]"
                            fontes['Autores'] = 'Semantic Scholar'
        # Abafa o erro
        except: pass

    # 3. Busca pela API do OpenAlex (ótimo para pesquisa acadêmica moderna)
    # Openalex funciona melhor batendo de frente com um DOI, então testamos se a gente precisa enriquecer e se já temos o DOI em mãos
    if precisa_enriquecer() and estado['doi']:
        # Tenta
        try:
            # Joga na API procurando pelo DOI específico daquele trabalho, então não tem erro de pegar artigo errado
            r = requests.get(f"https://api.openalex.org/works/https://doi.org/{estado['doi']}", timeout=5)
            # Confere a resposta
            if r.status_code == 200:
                # Transforma pra json
                it = r.json()
                # Se precisamos de resumo e a API do OpenAlex entregou a versão invertida e codificada dele
                if not estado['resumo'] and it.get('abstract_inverted_index'):
                    # Salva esse modelo doido de dados invertido
                    idx_alex = it['abstract_inverted_index']
                    # Um pequeno truquezinho de Python pra desinverter esse índice e fazer as palavras se ordenarem usando a posição real delas no texto original
                    pal = [(pos, p) for p, posicoes in idx_alex.items() for pos in posicoes]
                    # Organiza os tuplos baseando-se na primeira parte que é o número original
                    pal.sort(); 
                    # Junta só as palavras certinhas na sequência com espaços
                    estado['resumo'] = " ".join([p[1] for p in pal])
                    # Anota que o Openalex finalmente conseguiu ajudar a gente
                    fontes['Resumo'] = 'OpenAlex'
                
                # Se não temos autores, tenta extrair do OpenAlex
                if not estado['autores'] and it.get('authorships'):
                    lista_autores = [a.get('author', {}).get('display_name', '') for a in it.get('authorships', []) if a.get('author', {}).get('display_name')]
                    if lista_autores:
                        estado['autores'] = "; ".join(lista_autores)
                        if len(lista_autores) > 5:
                            estado['autores'] += " [Consórcio?]"
                        fontes['Autores'] = 'OpenAlex'
        # Se cair ignora
        except: pass

    # 4. Usar nosso trator Web Scraper (baixar os dados na marra da URL)
    # Define o nosso alvo principal, que é o Link do repositório, ou se ele for vazio a gente improvisa o link direto usando o site doi.org
    url_alvo = estado['link'] or (f"https://doi.org/{estado['doi']}" if estado['doi'] else None)
    # Se ainda tem que enriquecer e também a nossa url alvo for um link válido na web (que tenha a palavra http nela)
    if precisa_enriquecer() and url_alvo and url_alvo.startswith("http"):
        # Manda o link lá pra nossa função grandona e pega o dicionário com tudo o que ela tiver conseguido pescar de lá
        web_data = extrair_metadados_web_pdf(url_alvo, estado['titulo'], estado['autores'])
        # Se nós ainda estamos sem resumo e a nossa web achou
        if not estado['resumo'] and web_data['resumo']:
            # Pega o que foi achado
            estado['resumo'] = web_data['resumo']
            # Mantém anotada a fonte para gente avaliar se isso é confiável depois
            fontes['Resumo'] = web_data['fonte']
        # Faz o teste para título se precisava
        if not estado['titulo'] and web_data['titulo']:
            # Puxa pro nosso controle
            estado['titulo'] = web_data['titulo']
            # Marca o repositório
            fontes['Título'] = web_data['fonte']
        # Analisa os autores caso não tivessem e a web tenha puxado a lista
        if not estado['autores'] and web_data['autores']:
            # Atualiza
            estado['autores'] = web_data['autores']
            # Verifica a quantidade de autores para evitar consórcios gigantes
            if len(re.split(r';|\band\b', web_data['autores'])) > 5:
                estado['autores'] += " [Consórcio?]"
            # Registra
            fontes['Autores'] = web_data['fonte']
        # Vê de qual ano é
        if not estado['ano'] and web_data['ano']:
            # Salva
            estado['ano'] = web_data['ano']
            # Registra a fonte
            fontes['Ano'] = web_data['fonte']
        # Se a web conseguiu extrair o DOI correto e a gente ainda não tinha
        if not estado['doi'] and web_data.get('doi'):
            # Salva o novo DOI
            estado['doi'] = web_data['doi']
            # Registra a fonte
            fontes['DOI'] = web_data['fonte']

    # Uma última medida pra garantir o acesso: se nós não encontramos o link do repositório final de jeito nenhum até agora, mas a gente por um acaso tem o DOI salvo com a gente
    if not estado['link'] and estado['doi']:
        # Nós usamos o site do doi pra criar um link absoluto que resolve o artigo sozinho, então o pesquisador nunca fica a pé
        estado['link'] = f"https://doi.org/{estado['doi']}"
        # Marca nas fontes que fomos nós que geramos através do código
        fontes['Link'] = 'Gerado via DOI'

    # Imprime o resumo que nós coletamos bonitinho no painel
    imprimir_resumo_artigo(estado['titulo'], estado, fontes, idx, total)
    
    # Adiciona a informação de qual foi a fonte salvadora do resumo pra gente preencher aquela coluna do Excel na próxima etapa do processamento
    estado['fonte_resumo'] = fontes['Resumo']
    
    # Devolve o artigo que a gente mexeu de volta pro for preencher a nossa planilha em paz
    return estado

# ==========================================
# Funções de Auditoria Profunda 360 Graus
# ==========================================
def calcular_similaridade(str1, str2):
    if not str1 or not str2 or pd.isna(str1) or pd.isna(str2): return 0.0
    return difflib.SequenceMatcher(None, str(str1).lower(), str(str2).lower()).ratio()

def resgatar_doi_perdido(titulo, autores):
    if not titulo or titulo == 'N/A': return None, None, "Sem título válido para realizar a busca"
    titulo_limpo = urllib.parse.quote(titulo[:100]) # Limite para não quebrar APIs
    motivo_falha = "Artigo não indexado nos bancos de dados ou Google"
    
    # Tentativa 1: Crossref
    try:
        r = requests.get(f"https://api.crossref.org/works?query.title={titulo_limpo}&select=DOI,title,author,issued,URL&rows=1", timeout=5)
        if r.status_code == 200:
            dados_lista = r.json().get('message', {}).get('items', [])
            if dados_lista:
                dados = dados_lista[0]
                doi = dados.get('DOI')
                titulo_cr = dados.get('title', [''])[0]
                if calcular_similaridade(titulo, titulo_cr) > 0.85:
                    return doi, dados, ""
                else:
                    motivo_falha = f"Artigos encontrados tinham títulos muito diferentes (Achou: '{titulo_cr[:40]}...')"
    except: pass
    
    # Tentativa 2: Semantic Scholar
    try:
        r = requests.get(f"https://api.semanticscholar.org/graph/v1/paper/search?query={titulo_limpo}&limit=1&fields=externalIds,title,authors,year,openAccessPdf", timeout=5)
        if r.status_code == 200:
            dados_lista = r.json().get('data', [])
            if dados_lista:
                dados = dados_lista[0]
                doi = dados.get('externalIds', {}).get('DOI')
                titulo_ss = dados.get('title', '')
                if doi and calcular_similaridade(titulo, titulo_ss) > 0.85:
                    # Formata os dados pro padrão que nossas funções entendem
                    link_oa = dados.get('openAccessPdf', {}).get('url') if dados.get('openAccessPdf') else None
                    dados_formatados = {'author': [{'family': a.get('name', '').split()[-1]} for a in dados.get('authors', [])], 'issued': {'date-parts': [[dados.get('year')]]}, 'link_alternativo': link_oa}
                    return doi, dados_formatados, ""
                elif doi:
                    motivo_falha = f"Artigos encontrados tinham títulos muito diferentes (Achou: '{titulo_ss[:40]}...')"
    except: pass
    
    # Tentativa 3: OpenAlex
    try:
        r = requests.get(f"https://api.openalex.org/works?search={titulo_limpo}&per-page=1", timeout=5)
        if r.status_code == 200:
            dados_lista = r.json().get('results', [])
            if dados_lista:
                dados = dados_lista[0]
                doi_url = dados.get('doi')
                if doi_url: doi = doi_url.replace('https://doi.org/', '')
                else: doi = None
                titulo_oa = dados.get('title', '')
                if doi and calcular_similaridade(titulo, titulo_oa) > 0.85:
                    link_oa = dados.get('open_access', {}).get('oa_url') if dados.get('open_access', {}).get('is_oa') else None
                    dados_formatados = {'author': [{'family': a.get('author', {}).get('display_name', '').split()[-1]} for a in dados.get('authorships', [])], 'issued': {'date-parts': [[dados.get('publication_year')]]}, 'link_alternativo': link_oa}
                    return doi, dados_formatados, ""
                elif doi:
                    motivo_falha = f"Artigos encontrados tinham títulos muito diferentes (Achou: '{titulo_oa[:40]}...')"
    except: pass

    # Tentativa 4: Google Scholar (Busca bruta por Regex de DOI no HTML)
    time.sleep(3) # Pausa obrigatória anti-bloqueio
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'}
    try:
        r = requests.get(f"https://scholar.google.com/scholar?q={titulo_limpo}", headers=headers, timeout=10)
        if r.status_code == 200 and 'class="gs_ri"' in r.text:
            # Procura por um padrão clássico de DOI no texto da página de resultados
            match = re.search(r'\b(10\.\d{4,9}/[-._;()/:A-Za-z0-9]+)\b', r.text)
            if match:
                return match.group(1), None, ""
    except: pass

    # Tentativa 5: Google Normal (Busca bruta por Regex no HTML)
    time.sleep(3)
    try:
        r = requests.get(f"https://www.google.com/search?q={titulo_limpo}", headers=headers, timeout=10)
        if r.status_code == 200:
            match = re.search(r'\b(10\.\d{4,9}/[-._;()/:A-Za-z0-9]+)\b', r.text)
            if match:
                return match.group(1), None, ""
    except: pass

    return None, None, motivo_falha

def auditar_doi(doi, titulo):
    if not doi or doi == 'N/A': return "⚠️ Sem DOI (N/A)", None
    doi_limpo = urllib.parse.quote(doi)
    link_ref = f"https://doi.org/{doi}"
    
    # Tentativa 1: Crossref
    try:
        r = requests.get(f"https://api.crossref.org/works/{doi_limpo}", timeout=5)
        if r.status_code == 200:
            dados = r.json().get('message', {})
            titulo_cr = dados.get('title', [''])[0]
            if calcular_similaridade(titulo, titulo_cr) < 0.70: return f"⚠️ DOI Suspeito (Títulos não batem no Crossref. Ref: {link_ref})", dados
            return f"✅ DOI Confirmado via Crossref (Ref: {link_ref})", dados
    except: pass
    
    # Tentativa 2: Semantic Scholar
    try:
        r = requests.get(f"https://api.semanticscholar.org/graph/v1/paper/{doi_limpo}?fields=title,authors,year,url,openAccessPdf", timeout=5)
        if r.status_code == 200:
            dados = r.json()
            titulo_ss = dados.get('title', '')
            link_ref = dados.get('url', link_ref)
            link_oa = dados.get('openAccessPdf', {}).get('url') if dados.get('openAccessPdf') else None
            dados_formatados = {'author': [{'family': a.get('name', '').split()[-1]} for a in dados.get('authors', [])], 'issued': {'date-parts': [[dados.get('year')]]}, 'link_alternativo': link_oa}
            if calcular_similaridade(titulo, titulo_ss) < 0.70: return f"⚠️ DOI Suspeito (Títulos não batem no Semantic Scholar. Ref: {link_ref})", dados_formatados
            return f"✅ DOI Confirmado via Semantic Scholar (Ref: {link_ref})", dados_formatados
    except: pass
    
    # Tentativa 3: OpenAlex
    try:
        r = requests.get(f"https://api.openalex.org/works/https://doi.org/{doi_limpo}", timeout=5)
        if r.status_code == 200:
            dados = r.json()
            titulo_oa = dados.get('title', '')
            link_ref = dados.get('id', link_ref)
            link_oa = dados.get('open_access', {}).get('oa_url') if dados.get('open_access', {}).get('is_oa') else None
            dados_formatados = {'author': [{'family': a.get('author', {}).get('display_name', '').split()[-1]} for a in dados.get('authorships', [])], 'issued': {'date-parts': [[dados.get('publication_year')]]}, 'link_alternativo': link_oa}
            if calcular_similaridade(titulo, titulo_oa) < 0.70: return f"⚠️ DOI Suspeito (Títulos não batem no OpenAlex. Ref: {link_ref})", dados_formatados
            return f"✅ DOI Confirmado via OpenAlex (Ref: {link_ref})", dados_formatados
    except: pass
    
    # ÚLTIMO CASO: Tenta o Google Scholar direto na unha com muito cuidado
    time.sleep(3) # Dorme 3 segundos obrigatoriamente para o Google não bloquear de cara
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'}
        link_google = f"https://scholar.google.com/scholar?q={doi_limpo}"
        r = requests.get(link_google, headers=headers, timeout=10)
        
        if r.status_code == 429 or "Aguarde um momento" in r.text or "show you're not a robot" in r.text:
            return "⚠️ DOI Não Validado (Google Scholar Bloqueou IP por Excesso de Buscas)", None
            
        if r.status_code == 200:
            # Se ele encontrou a div de resultado clássica do Google Scholar, o artigo existe!
            if 'class="gs_ri"' in r.text:
                return f"✅ DOI Confirmado via Google Scholar (Ref: {link_google})", None
            else:
                return f"⚠️ DOI Não Encontrado em Nenhum Banco de Dados ou Google (Busca: {link_google})", None
    except:
        return "⚠️ Erro de Conexão na Validação do DOI", None

    return "⚠️ DOI Não Encontrado em Nenhum Banco de Dados", None

def auditar_autores(autores_zotero, dados_crossref):
    if not dados_crossref or 'author' not in dados_crossref: return "⚠️ Autores: API Sem Dados para Confirmar"
    if not autores_zotero or autores_zotero == 'N/A': return "⚠️ Autores Ausentes no Zotero"
    
    autores_oficiais = [a.get('family', '').lower() for a in dados_crossref['author'] if a.get('family')]
    zotero_lower = str(autores_zotero).lower()
    
    encontrados = sum(1 for oficial in autores_oficiais if oficial in zotero_lower)
    if autores_oficiais and encontrados == 0:
        return f"⚠️ Autores Suspeitos (Oficiais: {', '.join(autores_oficiais).title()})"
    return "✅ Autores Confirmados"

def auditar_ano(ano_zotero, dados_crossref):
    if not dados_crossref or 'issued' not in dados_crossref: return "⚠️ Ano: API Sem Dados para Confirmar"
    try:
        ano_oficial = str(dados_crossref['issued']['date-parts'][0][0])
        ano_zot = str(ano_zotero).strip()
        if ano_zot != 'N/A' and ano_zot != ano_oficial:
            return f"⚠️ Ano Suspeito (Zotero: {ano_zot} | Oficial: {ano_oficial})"
        return f"✅ Ano Confirmado ({ano_oficial})"
    except:
        pass
    return "⚠️ Ano: Falha ao Ler Dados Oficiais"

def extrair_link_alternativo(dados):
    if not dados: return None
    # Prioriza o campo padronizado nas funções acima
    if dados.get('link_alternativo'): return dados.get('link_alternativo')
    # Senão tenta no formato do Crossref nativo
    if 'link' in dados:
        for l in dados['link']:
            if l.get('intended-application') == 'text-mining' and l.get('URL'):
                return l['URL']
    return None

def auditar_link(link, link_alternativo=None):
    if not link or link == 'N/A': return "⚠️ Link Ausente"
    try:
        r = requests.head(link, timeout=5, allow_redirects=True, verify=False)
        if r.status_code >= 400:
            # Se tomou Bot Protection (403/401) e tem um PDF de Open Access
            if r.status_code in [403, 401] and link_alternativo:
                r_alt = requests.head(link_alternativo, timeout=5, allow_redirects=True, verify=False)
                if r_alt.status_code < 400:
                    return f"✅ Link Oficial Protegido Anti-Bot, mas PDF Gratuito Encontrado ({link_alternativo})"
                    
            if r.status_code not in [403, 401, 405]:
                r2 = requests.get(link, timeout=5, stream=True, verify=False)
                if r2.status_code >= 400 and r2.status_code not in [403, 401]:
                    if link_alternativo: return f"⚠️ Link Quebrado (Erro {r2.status_code}). Alternativo Disponível: {link_alternativo}"
                    return f"⚠️ Link Quebrado (Erro {r2.status_code} na url {link})"
            
            # Se chegou aqui é porque era só proteção de bot e não temos/não funcionou o alternativo
            if link_alternativo: return f"✅ Link Oficial Protegido Anti-Bot (Tente o Alternativo Livre: {link_alternativo})"
            return f"✅ Link Ativo porém Protegido Anti-Bot ({link})"
            
        return f"✅ Link Ativo ({link})"
    except requests.exceptions.Timeout:
        if link_alternativo: return f"⚠️ Link Oficial Timeout. Tente o Alternativo Livre: {link_alternativo}"
        return f"⚠️ Link Quebrado (Timeout na url {link})"
    except Exception:
        if link_alternativo: return f"⚠️ Link Oficial Falhou. Tente o Alternativo Livre: {link_alternativo}"
        return f"⚠️ Link Quebrado (Falha de Conexão na url {link})"

def auditar_resumo(resumo, titulo):
    if not resumo or resumo == 'N/A': return "⚠️ Sem Resumo"
    resumo_str = str(resumo).lower()
    avisos = []
    
    if len(resumo_str) < 150:
        avisos.append("⚠️ Resumo Muito Curto")
        
    blacklist = ['enable javascript', 'access denied', 'cookies policy', 'login required', 'please log in', 'robot', 'captcha', 'security check']
    if any(b in resumo_str for b in blacklist):
        avisos.append("⚠️ Resumo Inválido (Scraper de Erro)")
        
    whitelist = ['in this paper', 'we propose', 'the aim', 'this study', 'this research', 'we introduce', 'abstract', 'neste artigo', 'este estudo']
    tem_marcador = any(w in resumo_str for w in whitelist)
    
    titulo_limpo = re.sub(r'[^\w\s]', '', str(titulo).lower())
    palavras_titulo = [p for p in titulo_limpo.split() if len(p) > 4]
    
    palavras_encontradas = sum(1 for p in palavras_titulo if p in resumo_str)
    if palavras_titulo and palavras_encontradas == 0:
        avisos.append("⚠️ Resumo Suspeito (Falta de Palavras do Título)")
        
    if not avisos and tem_marcador:
        return "✅ Resumo Confirmado (Alta Confiança)"
    elif not avisos:
        return "✅ Resumo Aparentemente Válido"
    else:
        return " | ".join(avisos)

def auditar_item(item):
    avisos = []
    novo_doi = None
    dados_oficiais = None
    
    # 1. Valida o DOI
    if item['DOI'] == 'N/A':
        resgate, dados, motivo = resgatar_doi_perdido(item['TÍTULO'], item['AUTORES'])
        if resgate:
            novo_doi = resgate
            dados_oficiais = dados
            avisos.append(f"✅ DOI Resgatado (https://doi.org/{novo_doi})")
        else:
            avisos.append(f"❌ DOI Ausente Absoluto ({motivo})")
    else:
        status_doi, dados = auditar_doi(item['DOI'], item['TÍTULO'])
        dados_oficiais = dados
        avisos.append(status_doi)
        
    # 2. Valida o Resumo
    avisos.append(auditar_resumo(item['RESUMO'], item['TÍTULO']))
        
    # 3. Valida Autores
    avisos.append(auditar_autores(item['AUTORES'], dados_oficiais))
        
    # 4. Valida Ano de Publicação
    avisos.append(auditar_ano(item['ANO DE PUBLICAÇÃO'], dados_oficiais))
        
    # 5. Valida Link
    link_para_testar = f"https://doi.org/{novo_doi}" if novo_doi else item['LINK']
    link_alternativo = extrair_link_alternativo(dados_oficiais)
    avisos.append(auditar_link(link_para_testar, link_alternativo))
        
    return avisos, novo_doi

# ==========================================
# Início do processamento principal e do gerador do Excel
# ==========================================
# Imprime os detalhes pro usuário se situar na tela preta
print("======================================================")
print(" 🚀 Fazendo a Revisão Sistemática da Literatura All_Dataset ")
print("======================================================")
# Carrega da nuvem do Zotero toda a lista completa da nossa biblioteca que fica no topo (sem contar notas soltas ou pdfs anexos por fora) e forma uma mega lista em memória
items = [i for i in zot.everything(zot.top()) if i['data']['itemType'] not in ['attachment', 'note']]
# Confere o tamanho dessa lista pra gente conseguir dizer se estamos em um terço ou na metade depois
total = len(items)

# Inicializa um vetor onde as linhas da tabela final do nosso dataset completo vão se agrupar
dados_all_dataset = []

# Verifica se existe pelo menos um item que possua ranking manual válido preenchido no Zotero
tem_ranking_manual = any(re.search(r'^\s*\d+(?:\s|\n|$)', str(i.get('data', {}).get('extra', ''))) for i in items)

# Varre os itens vindos da biblioteca um a um numerados desde o número 1 graças a função enumerate
for idx, item in enumerate(items, 1):
    # Foca nos dados brutos do item da vez
    d = item['data']
    # Dispara todo o nosso motor de caçador de inteligência pra enriquecer os vazios que estão neste artigo e recebe um arquivo já vitaminado de volta
    estado_enriquecido = buscar_metadados_hierarquia(item, idx, total, tem_ranking_manual)
    
    # Prepara o link que vai direto pro artigo. Pega o estado enriquecido. Mas se por algum milagre ele ainda estiver vazio de link tenta criar a URL do DOI, e se nada mais der certo mete o triste N/A
    link_final = estado_enriquecido['link'] if estado_enriquecido['link'] else (f"https://doi.org/{estado_enriquecido['doi']}" if estado_enriquecido['doi'] else 'N/A')
    
    # Adiciona na nossa tabela, organizando linha por linha para criar colunas mais tarde, caso um elemento continue vazio preenchemos a falha com o texto N/A
    dados_all_dataset.append({
        # Grava a coluna do Ranking com zero à esquerda
        'RANK': estado_enriquecido['rank'],
        # Grava a coluna do DOI
        'DOI': estado_enriquecido['doi'] if estado_enriquecido['doi'] else 'N/A',
        # Grava os dados dos autores e pesquisadores
        'AUTORES': estado_enriquecido['autores'] if estado_enriquecido['autores'] else 'N/A',
        # Grava o título final da publicação
        'TÍTULO': estado_enriquecido['titulo'] if estado_enriquecido['titulo'] else 'N/A',
        # Grava o resumo grande que nós pegamos do zotero ou caçamos em quatro camadas
        'RESUMO': estado_enriquecido['resumo'] if estado_enriquecido['resumo'] else 'N/A',
        # O ano exato da publicação para manter linha do tempo
        'ANO DE PUBLICAÇÃO': estado_enriquecido['ano'] if estado_enriquecido['ano'] else 'N/A',
        # Fixando a base de dados como Google Scholar, conforme solicitado, ignorando os metadados brutos do Zotero que podem vir sujos
        'BASE DE DADOS': 'Google Scholar',
        # Finca o link real gerado pro excel
        'LINK': link_final
    })
    # Congela o loop por exato um segundo pra gente não tomar flag de robô frenético em nenhuma das APIs, respeitando os termos mundiais (Polite)
    time.sleep(1)

# Ponto final de coletas na web. Daqui em diante, podemos rodar nossa auditoria antes de gerar o excel.
print("\n======================================================")
resposta = input("🔎 Deseja rodar a Auditoria Profunda de DOIs e Resumos? (Isso pode demorar alguns minutos) [S/N]: ").strip().upper()
if resposta == 'S':
    print("🕵️ Iniciando Auditoria Profunda. Por favor, aguarde...")
    for i, item in enumerate(dados_all_dataset):
        sys.stdout.write(f"\rAuditando item {i+1}/{len(dados_all_dataset)}...")
        sys.stdout.flush()
        avisos, novo_doi = auditar_item(item)
        if novo_doi:
            item['DOI'] = novo_doi
            item['LINK'] = f"https://doi.org/{novo_doi}"
        item['AVISOS DE AUDITORIA'] = " | ".join(avisos) if avisos else "OK"
        # Aguarda meio segundo para não sobrecarregar as APIs de validação e não sermos bloqueados
        time.sleep(0.5)
    print("\n✅ Auditoria concluída com sucesso!")
else:
    for item in dados_all_dataset:
        item['AVISOS DE AUDITORIA'] = "N/A (Auditoria não executada)"

print("\n📊 Construindo Planilha...")
# Executa a função global que cria uma representação de planilha em branco
wb = Workbook()
# Tira a aba em branco 'Sheet' que já vem criada por padrão, vamos adicionar abas customizadas
wb.remove(wb.active)

# Ordena toda a nossa lista de forma crescente baseada no conteúdo da coluna RANK
dados_all_dataset = sorted(dados_all_dataset, key=lambda x: x['RANK'])

# Prepara a sequência das colunas de identificação mais básicas que nós coletamos (com o RANK na Coluna A)
base_cols = ['RANK', 'DOI', 'AUTORES', 'TÍTULO', 'RESUMO', 'ANO DE PUBLICAÇÃO', 'BASE DE DADOS', 'LINK', 'AVISOS DE AUDITORIA']
# As colunas da etapa de screening de título além de tudo da base tem mais dois atributos para marcarmos na mão dentro do excel
cols_scr = base_cols + ['STATUS (Incluído/Excluído)', 'MOTIVO EXCLUSÃO']
# E as colunas finais precisam de áreas de análise completas para nós completarmos o quadro PRISMA do documento científico
cols_fin = base_cols + ['TIPO DE ESTUDO', 'MÉTODO DE PESQUISA', 'RESULTADOS PRINCIPAIS', 'RELAÇÃO COM BRAINNODE', 'TECNOLOGIA UTILIZADA']

# Monta uma configuração super estruturada de quais abas vão ser criadas na planilha indicando nome, preenchimento e lista de colunas respectiva
abas_config = [
    # A base de tudo onde colamos nossos dados automatizados do script Python
    ("All_Dataset", dados_all_dataset, base_cols),
    # A aba onde os pesquisadores vão colocar o recorte do que sobrou limpando teses sem impacto e literaturas cinzas
    ("Remoção da Literatura Cinza", [], base_cols),
    # Depois o próprio ser humano remove ali os estudos idênticos duplicados
    ("Remoção de estudos duplicados", [], base_cols),
    # O funil se fecha pro screening só através dos títulos
    ("1 - Busca (Screening) - Título", [], cols_scr),
    # E depois lê integralmente pra dar check e cruzar com os motivos de exclusão
    ("2 - Leitura do Texto completo", [], cols_scr),
    # E sobram os selecionados finos, os chamados finalistas, na aba completa com todos os insights exigidos
    ("Estudos Finais", [], cols_fin)
]

# Configurações globais dos estilos, determinando o fundo que tem um degradê falso para parecer azul clássico corporativo em formato sólido
fill_azul = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
# Estabelece um traço contínuo fino para aplicar depois contornando os quatro lados de uma célula 
borda = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Começa a montar fisicamente a planilha percorrendo nossos esquemas de abas indicados lá em cima
for nome, dados, colunas in abas_config:
    # Insere uma aba do excel novinha com a etiqueta baseada no nome da vez
    ws = wb.create_sheet(nome)
    # Forma a matriz inteira do dataframe. Se for na aba inicial ele enche de dados capturados, caso for das abas da frente do fluxo ele monta com o esqueleto em branco.
    df = pd.DataFrame(dados, columns=colunas) if dados else pd.DataFrame(columns=colunas)
    
    # Transfere os dados (já com o rótulo do cabeçalho) puxados do nosso dataframe pros itens do Excel do modo correto sem quebrar por index
    for r in dataframe_to_rows(df, index=False, header=True): 
        # Insere fisicamente a linha gerada dentro dessa aba especificada na variável ws
        ws.append(r)
        
    # Quando acabar de puxar, volta e foca somente na linha 1 da aba, que é obrigatoriamente a do cabeçalho de títulos
    for cell in ws[1]: 
        # Dá de brinde aquele preenchimento azul corporativo do excel bonitão pra célula em questão
        cell.fill = fill_azul
        # Coloca as letrinhas da célula como cor branca e espessa (em formato bold)
        cell.font = Font(color="FFFFFF", bold=True)
        # O alinhamento centralizado garante a estética do visual, e a opção do texto descer ajuda caso a coluna acabe antes da letra do título
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Contorna todas as pontas da célula com linhas desenhadas
        cell.border = borda
        
    # Começa um segundo laço, dessa vez a partir da linha 2 até a última da aba para podermos cuidar dos dados das células em comum
    for row in ws.iter_rows(min_row=2):
        # Repassa célula por célula da linha
        for cell in row: 
            # A gente estipula a orientação vertical jogada para cima com o wrap ligado para os parágrafos se adaptarem descendo pelo tamanho de linhas, excelente pros resumos longos
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            # Adiciona a formatação padrão daquelas caixinhas nos arredores para construir um formato grade
            cell.border = borda
            
    # Ajusta as proporções da tela com a largura ideal percorrendo nosso mapa atual de cabeçalhos
    for c_idx, col_name in enumerate(colunas, 1):
        # Transforma o número que o laço nos dá para a letra compatível no cabeçalho do programa excel (ex: 1 vira A, 2 vira B)
        letra = ws.cell(row=1, column=c_idx).column_letter
        # Define os tamanhos: grandes áreas de texto levam 50 de largura; informações de média prioridade como nomes levam 30; dados pontuais (data, bases curtas) ganham 18 pra equilibrar a tela 
        ws.column_dimensions[letra].width = 50 if col_name in ['TÍTULO', 'RESUMO', 'RESULTADOS PRINCIPAIS', 'AVISOS DE AUDITORIA'] else (30 if col_name in ['AUTORES', 'LINK'] else 18)
        
    # Estipula um pino invisível embaixo do bloco A1. Isso ativa a função sensacional do congelamento de painéis, travando nosso cabeçalho lindamente enquanto rola pra baixo
    ws.freeze_panes = "A2"

# Define o nome base e a extensão do arquivo
nome_formatado = nome_usuario.strip() if nome_usuario.strip() else "Usuario"
nome_base = f"RSL_Dashboard_{nome_formatado}"
extensao = ".xlsx"
nome_arquivo = f"{nome_base}{extensao}"
contador = 1

# Fica testando se o arquivo já existe na pasta, se sim, vai somando +1 no nome pra não sobrescrever o antigo
while os.path.exists(nome_arquivo):
    nome_arquivo = f"{nome_base}_{contador}{extensao}"
    contador += 1

# Manda salvar o documento virtual recém-nascido fisicamente lá na mesma pasta de onde o python iniciou
wb.save(nome_arquivo)
# Coloca mensagem para encerrar o ciclo do usuário feliz e vitorioso
print(f"\n✅ Planilha '{nome_arquivo}' gerada com sucesso!")